import logging
import inspect
import csv
from openpyxl import workbook,load_workbook

class utils():
    def customer_logging(logLevel=logging.DEBUG):
        # logger_name = inspect.stack()[1][1]  # it will update loglevel
        logger = logging.getLogger("__name__")  #result will mention only with name
        # logger = logging.getLogger(logger_name)   # result will give defind loglevel
        logger.setLevel(logLevel)
        fh = logging.FileHandler("C:\\pythonProject1\\project_pytest\\report\\Yatra.log")
        formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(name)s - %(message)s ")
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger
    def read_excel_data(file_name,sheet):
        datalist=[]
        wb=load_workbook(filename=file_name)
        sh=wb[sheet]
        row_ct =sh.max_row
        col_ct =sh.max_column
        for i in range (2 ,row_ct+1):  # 1 row skip beacsue header is there
            row =[]
            for j in range(1,col_ct+1):
                row.append(sh.cell(row=i,column = j).value)
            datalist.append(row)
            print(datalist)
        return datalist

    def read_csv_data(filename):
        # datalist= []
        # csvdata = open(filename,"r")
        # reader =csv.reader(csvdata)
        # next(reader)
        # for row in reader:
        #     datalist.append(row)
        # return datalist

        with open(filename, 'r') as f:
            datalist =[]
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                datalist.append(row)
            return datalist